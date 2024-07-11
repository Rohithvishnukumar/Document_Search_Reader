# ------------------------------------Doc Search Reader-------------------------------------------
# ///////////////////////////////// Developed by Rohith Vishnu ////////////////////////////////////




from threading import Timer
# import pdfplumber
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from flask_mysqldb import MySQL
import os ,random
from flask import Flask, render_template, redirect, request, send_file, session
# from flask_session import Session
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
import fitz  # PyMuPDF
from docx import Document
import openpyxl
import pytesseract
from googletrans import Translator
from PIL import Image
import io
import re
from openpyxl.drawing.image import Image as OpenpyxlImage
import spacy
import nltk
import fitz
import pytesseract
import pkg_resources
from symspellpy import SymSpell, Verbosity
import string
from passlib.hash import pbkdf2_sha256
# import requests  # used for making HTTP server side requests
import google.generativeai as genai


nltk.download("wordnet")
from nltk.corpus import wordnet

nlp = spacy.load("en_core_web_md")  # pip install en_core_web_md
from dotenv import load_dotenv

load_dotenv()


pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"  # replaced with actual server directory
langvar = "eng+tel+afr+fra+ita+chi_sim+chi_tra+rus+hin+tam+san+kor+jpn+nld+vie+script_latin+deu+sqi+eus"


app = Flask(__name__)
mysql = MySQL(app)
mail = Mail(app)
app.secret_key = os.getenv("secret_key")


# Create uploads folder if it doesn't exist
if not os.path.exists("./uploads"):
    os.mkdir("./uploads")
app.config["UPLOAD_FOLDER"] = "./uploads"


#--------------------MAIL Service for MFA----------------------

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')

mail = Mail(app)


# --------------- Database Connection (MySQL)  --------------------

app.config["MYSQL_HOST"] = os.getenv("MYSQL_HOST")
app.config["MYSQL_USER"] = os.getenv("MYSQL_USER")
app.config["MYSQL_PASSWORD"] = os.getenv("MYSQL_PASSWORD")
app.config["MYSQL_DB"] = os.getenv("MYSQL_DB")


@app.route("/", methods=["GET"])
def index():
    return render_template("login.html")

@app.route("/return", methods=["GET"])
def routereturn():
    return render_template("index.html", username = session["username"])



@app.route("/submit", methods=["GET", "POST"])
def submit():
    if request.method == "POST":
        if "inpfile" not in request.files:
            return "An Error Occurred - File not Found"

        file = request.files["inpfile"]
        if file.filename == "":
            return "File name is NULL"

        if file:
             
            filename = secure_filename(file.filename)
            session["filename"] = filename
            file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            file.save(file_path)
            # Process the uploaded file
            processed_text = process_document(file_path, filename)

            return processed_text

    return "General Return"


def process_document(file_path, filename):
    if filename.endswith(".pdf"):
        text = extract_text_from_pdf(file_path)
    elif filename.endswith(".docx"):
        text = extract_text_from_docx(file_path)
    elif filename.endswith(".xlsx"):
        text = extract_text_from_excel(file_path)
    elif filename.lower().endswith((".png", ".jpg", ".jpeg")):
        text = extract_text_from_image(file_path)
    else:
        errdata = "Unsupported file format"
        return render_template("error_page.html", errdata=errdata)

    # text = translate_to_english(text)
    # text = "Rohith <span>Vishnu</span>"
    # return render_template("trans.html", text = text)
    return spellCheck(text)


# --------------------------------- Text extraction and OCR-----------------------------------


# def clean_text(text):
#     # Replace newlines with spaces
#     text = text.replace("\n", " ")
#     # Remove multiple spaces
#     text = re.sub(r"\s+", " ", text)
#     # Example of fixing common OCR splitting issues, this can be expanded as needed
#     text = re.sub(r"\bcust omer\b", "customer", text, flags=re.IGNORECASE)
#     # Add other specific text cleaning steps as needed
#     return text






def clean_text(text):
    # Replace multiple newlines with a single newline
    text = re.sub(r"[\r\n]+", "\n", text)
    # Remove leading and trailing whitespace
    text = text.strip()
    # Example of fixing common OCR splitting issues, this can be expanded as needed
    text = re.sub(r"\bcust omer\b", "customer", text, flags=re.IGNORECASE)
    # Replace any remaining newlines with spaces (optional, if needed)
    text = text.replace("\n", " ")
    return text



def extract_text_from_pdf(file_path):
    text = ""
    doc = fitz.open(file_path)

    for page_num in range(len(doc)):
        page = doc[page_num]
        # Extract text from the page
        try:
            page_text = page.get_text()
            text += page_text + "\n"
        except Exception as e:
            print(f"Error extracting text from page {page_num}: {e}")

        # Extract images from the page
        image_list = page.get_images(full=True)
        for img_index, img in enumerate(image_list):
            try:
                xref = img[0]
                base_image = doc.extract_image(xref)
                image_bytes = base_image["image"]
                image = Image.open(io.BytesIO(image_bytes))
                image_text = pytesseract.image_to_string(image, lang=langvar)
                text += image_text + "\n"
            except Exception as e:
                print(
                    f"Error extracting text from image {img_index} on page {page_num}: {e}"
                )

    # Clean the extracted text
    cleaned_text = clean_text(text)
    return cleaned_text


# def extract_text_from_pdf(file_path):
#     text = ""
#     doc = fitz.open(file_path)
#     for page in doc:
#         text += page.get_text() + "\n"

#         # Extract images from the page
#         image_list = page.get_images(full=True)
#         # print(list(enumerate(image_list)))
#         for img_index, img in enumerate(image_list):
#             xref = img[0]

#             base_image = doc.extract_image(xref)
#             image_bytes = base_image["image"]
#             image = Image.open(io.BytesIO(image_bytes))
#             text += pytesseract.image_to_string(image, lang=langvar)

#     return text




def extract_text_from_docx(file_path):
    text = ""
    doc = Document(file_path)
    for para in doc.paragraphs:
        text += para.text + "\n\n"

    # Extract images from the document
    for rel in doc.part.rels:
        if "image" in doc.part.rels[rel].target_ref:
            img = doc.part.rels[rel].target_part.blob
            image = Image.open(io.BytesIO(img))
            text += pytesseract.image_to_string(image, lang=langvar)

    return clean_text(text)


def extract_text_from_excel(file_path):
    text = ""
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text += f"Sheet: {sheet_name}\n"
        for row in sheet.iter_rows():
            row_text = " | ".join(
                [str(cell.value) if cell.value else "" for cell in row]
            )
            text += row_text + "\n"

        # Extract images from the sheet
        for image in sheet._images:
            img_stream = image.ref
            img_stream.seek(0)
            img = Image.open(io.BytesIO(img_stream.read()))
            text += pytesseract.image_to_string(img, lang=langvar) + "\n"

    return text


def extract_text_from_image(file_path):
    text = pytesseract.image_to_string(Image.open(file_path), lang=langvar)
    return text


def translate_to_english(text):
    translator = Translator()
    translated = translator.translate(text, src="auto", dest="en")
    return translated.text


# --------------------------- Spell Check  --------------------------------------


def spellCheck(text):
    sym_spell = SymSpell(max_dictionary_edit_distance=3, prefix_length=7)
    dictionary_path = pkg_resources.resource_filename(
        "symspellpy", "frequency_dictionary_en_82_765.txt"
    )
    sym_spell.load_dictionary(dictionary_path, term_index=0, count_index=1)

    input_terms = separate_words(text)  # Separate words from text

    for input_term in input_terms:
        suggestions = sym_spell.lookup(
            input_term,
            Verbosity.CLOSEST,
            max_edit_distance=2,
            transfer_casing=True,
            include_unknown=True,
        )

        if len(suggestions) > 0 and input_term != suggestions[0].term:

            reptext = f"<span>{input_term}</span>"
            text = text.replace(input_term, reptext)

            # if input_term not in spellDic.values():
            #     if spellDic:
            #         last_key = max(spellDic.keys()) + 1
            #     else:
            #         last_key = 0

            #     spellDic[last_key] = input_term

        # print(spellDic)

    global textglobal
    textglobal = text
    return render_template("trans.html", text=text)


def separate_words(text):
    # Remove punctuation (including apostrophes)
    translator = str.maketrans("", "", string.punctuation)
    clean_text = text.translate(translator)

    # Split the text into words
    words = clean_text.split()
    return words


#  -----------------------------------------Search-------------------------------------------------


def search_keywords_and_synonyms_and_phrases(text, keywords):
    keywords = keywords.split(",")
    all_keywords = set()
    doc = nlp(text.lower())

    # Expand with synonyms
    for keyword in keywords:
        synonyms = get_synonyms(keyword)
        all_keywords.add(keyword.strip().lower())
        all_keywords.update(synonym.strip().lower() for synonym in synonyms)

    # Highlight keywords, synonyms, and similar phrases
    highlighted_text = text
    for word in all_keywords:
        highlighted_text = re.sub(
            r"(?<!\w)({})(?!\w)".format(re.escape(word)),
            r"<mark>\1</mark>",
            highlighted_text,
            flags=re.IGNORECASE,
        )

    # Highlight similar phrases
    for sent in doc.sents:
        for keyword in all_keywords:
            if nlp(keyword).similarity(nlp(sent.text.lower())) > 0.7:
                highlighted_text = highlighted_text.replace(
                    sent.text, f"<mark>{sent.text}</mark>"
                )

    return highlighted_text


def get_synonyms(word):
    synonyms = set()
    for syn in wordnet.synsets(word):
        for lemma in syn.lemmas():
            synonyms.add(lemma.name().replace("_", " "))
    return synonyms


@app.route("/search", methods=["GET", "POST"])
def search():
    keywords = request.form["srch"]
    highlighted_text = search_keywords_and_synonyms_and_phrases(textglobal, keywords)
    return render_template("trans.html", text=highlighted_text)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")
    
    if(request.method == "POST"):

        session["username"] = request.form["username"]
        session["password"] = request.form["password"]

        usrlog = session["username"]
        pwdlog = session["password"]

        if(len(usrlog) < 4 ):
            return render_template("error_page.html")

        try:
            hashlog = pbkdf2_sha256.hash(pwdlog)
            cur = mysql.connection.cursor()
            cur.execute(f"SELECT password FROM users WHERE username = '{usrlog}' ")
            result = cur.fetchall()[0][0]

        except:
            return render_template("login.html", msg = "User Name or Password does not Match Try Again Fatal")
        
    
        if( pbkdf2_sha256.verify(pwdlog , result) ):

            cur.execute(f"SELECT userid FROM users WHERE username = '{usrlog}' ")
            session['userid'] = cur.fetchall()[0][0]
            print(session['userid'])

            
            rand = random.random() * 1000000
            session['otp'] = int(rand)
            session['mfa_enabled'] = True
            otp = session['otp']

            cur = mysql.connection.cursor()
            cur.execute(f"SELECT recovery FROM users WHERE username = '{session['username']}' ")
            recov = cur.fetchall()[0][0]
            print(recov)

            msg = Message('Doc Search Reader', sender ='rohithvishnuk@gmail.com', recipients = [recov] ) 
            msg.body = f'Your OTP request for Doc Search Reader is: {otp} \n\n Developed by Rohith Vishnu'
            mail.send(msg) 

           
            return render_template("MFA.html")
        
        else:
            return render_template("login.html", msg = "User Name or Password does not Match Try Again")



@app.route("/mfa", methods = [ "GET",  "POST"])
def mfa():

    if(request.method == "GET"):
        return render_template("login.html")

    otpcheck = int(request.form["numotp"])

    if(otpcheck == session['otp']):
        # session.pop("otp")
        return render_template("index.html", username = session["username"])
    
    else:
        # session.pop("otp")
        return render_template("login.html", msg = "Your OTP is incorrect. Try Logging in again" )




@app.route("/signup", methods=["GET", "POST"])
def signup():

    if request.method == "GET":
        return render_template("signup.html")

    if request.method == "POST":

        usr = request.form["username"]
        pwd = request.form["password"]
        recov = request.form["email"]

        passpattern = r'''^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$'''

        veri = re.match(passpattern,pwd)

        if(len(usr) < 4 or veri == False ):
            return render_template("error_page.html" , errdata = "Username or Password criteria does not match (Server side Validation) client side validation overridden")

        hash = pbkdf2_sha256.hash(pwd)


        cur = mysql.connection.cursor()
        cur.execute("SELECT username FROM users")
        result = cur.fetchall()
        for i in result:
            for j in i:
                if usr == j:
                    exist = "Username Already Exist. Try With a different User Name"
                    cur.close()
                    return render_template("signup.html", exist=exist)

        cur.execute("INSERT INTO users(username, password, recovery) VALUES(%s, %s, %s)", (usr, hash, recov))
        mysql.connection.commit()
        cur.close()
        
        return redirect("/login")



@app.route("/summarypage", methods = ["GET","POST"])
def summaryroute():

    if(request.method == "POST"):

        file = "Summary_" + session["filename"]
        
        cur = mysql.connection.cursor()

        data = request.get_json()
        sumtext = data['text']

        usr = int(session["userid"])

        cur.execute(f"SELECT filename FROM summary WHERE userid = {usr}")
        res = cur.fetchall()

        for i in res:
            for j in i:
                if(j == file):
                    file = file + "(1)" 

        cur.execute('INSERT INTO summary VALUES(%s, %s, %s )', ( usr , file, sumtext ))
        mysql.connection.commit()
        print("successfully Inserted")


        print("Redirecting to records")
        return redirect("/records")





@app.route("/records", methods = ["GET", "POST"])
def records():

    if(request.method == "GET"):

        usr = session["username"]
        userid = int(session["userid"])

        cur = mysql.connection.cursor()
        cur.execute(f"SELECT * FROM summary WHERE userid = {userid} ")
        alldata = cur.fetchall()
        cur.close()

        data = {
            "table" : alldata,
            "usrname" : usr,
            "test" : "Hello"
        }

        return render_template("records.html", data = data)



def wrap_text(text, max_width, font_size, canvas):
    lines = []
    line = ""
    words = text.split()
    
    for word in words:
        if canvas.stringWidth(line + " " + word, "Helvetica", font_size) < max_width:
            line += " " + word
        else:
            lines.append(line.strip())
            line = word
    
    if line:
        lines.append(line.strip())
    
    return lines




@app.route("/download/<file>")
def download_file(file):

    session["downfile"] = file

    c = canvas.Canvas(f"{file}.pdf", pagesize = A4)

    cur = mysql.connection.cursor()
    cur.execute(f"SELECT content FROM summary WHERE filename = '{file}' ")
    txt = cur.fetchall()[0][0]
    cur.close()

    c = canvas.Canvas(f"{file}.pdf", pagesize=A4)
    c.setFont("Helvetica", 12)

    max_width = A4[0] - 100  # Adjust margin
    lines = wrap_text(txt, max_width, 12, c)

    y_position = A4[1] - 50  # Start from top (A4[1]) minus margin
    for line in lines:
        c.drawString(50, y_position, line)
        y_position -= 15  # Move down 15 units for the next line

    c.showPage()
    c.save()

    t = Timer(10.0, delete_file, args=[f"{file}.pdf"])   # if file deletion is messing up in concurrent environments when you deploy comment it  
    t.start()

    return send_file(f"./{file}.pdf" , as_attachment= True)


def delete_file(pdf_filename):
    try:
        os.remove(pdf_filename)
    except Exception as e:
        print(f"Error deleting file: {e}")




if __name__ == "__main__":
    app.run(debug=True)
